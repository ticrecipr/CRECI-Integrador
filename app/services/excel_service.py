from pathlib import Path
from shutil import copy2
from datetime import datetime

from openpyxl import load_workbook

from app.core.constants import COLUNAS_PLANILHA


class ExcelService:

    def __init__(self, arquivo: str):

        self.arquivo = Path(arquivo)

        self.workbook = None

        self.sheet = None

        self.colunas = {}

    def criar_backup(self):

        backup_dir = Path("backup")

        backup_dir.mkdir(exist_ok=True)

        nome = datetime.now().strftime("%Y%m%d_%H%M%S")

        destino = backup_dir / f"{self.arquivo.stem}_{nome}.xlsx"

        copy2(self.arquivo, destino)

        return destino

    def abrir(self):

        self.workbook = load_workbook(self.arquivo)

    def listar_abas(self):

        return self.workbook.sheetnames

    def selecionar_aba(self, nome):

        self.sheet = self.workbook[nome]

    def localizar_colunas(self):

        self.colunas.clear()

        cabecalho = next(self.sheet.iter_rows(min_row=1, max_row=1))

        for celula in cabecalho:

            if celula.value is None:
                continue

            texto = str(celula.value).strip()

            self.colunas[texto] = celula.column

        obrigatorias = [

            COLUNAS_PLANILHA["cpf"],

            COLUNAS_PLANILHA["situacao"],

            COLUNAS_PLANILHA["financeira"],

            COLUNAS_PLANILHA["detalhe"]

        ]

        faltando = []

        for coluna in obrigatorias:

            if coluna not in self.colunas:

                faltando.append(coluna)

        if faltando:

            raise Exception(
                f"Colunas não encontradas: {', '.join(faltando)}"
            )

        return self.colunas